"""
Excel and CSV export functionality for EOC Roster Generator.
Generates beautifully formatted roster files.
"""

import csv
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter

from roster.models import Roster, ShiftType, DayAssignment


class RosterExporter:
    """Exports roster data to Excel and CSV formats."""
    
    # Color definitions
    COLORS = {
        'header_bg': 'FF1E3A5F',      # Dark blue
        'header_font': 'FFFFFFFF',     # White
        'lead_bg': 'FF90EE90',         # Light green
        'backup_bg': 'FFFFFFA0',       # Light yellow
        'weekend_bg': 'FFE6F2FF',      # Light blue
        'off_bg': 'FFD3D3D3',          # Light gray
        'shift1_header': 'FF4CAF50',   # Green
        'shift2_header': 'FF2196F3',   # Blue
        'shift3_header': 'FF9C27B0',   # Purple
        'border': 'FF000000',          # Black
    }
    
    def __init__(self, output_dir: str = "output"):
        """Initialize exporter with output directory."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Define styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _get_filename(self, roster: Roster, extension: str) -> str:
        """Generate filename for the roster."""
        month_name = roster.get_month_name()
        return f"EOC_Roster_{month_name}_{roster.year}.{extension}"
    
    def export_to_excel(self, roster: Roster) -> str:
        """Export roster to Excel file with formatting."""
        wb = Workbook()
        
        # Sheet 1: Monthly Calendar View
        self._create_calendar_sheet(wb, roster)
        
        # Sheet 2: Individual Schedules
        self._create_individual_sheet(wb, roster)
        
        # Sheet 3: Statistics
        self._create_statistics_sheet(wb, roster)
        
        # Remove default sheet if exists
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Save file
        filename = self._get_filename(roster, "xlsx")
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        return str(filepath)
    
    def _create_calendar_sheet(self, wb: Workbook, roster: Roster):
        """Create weekly roster with 3 shift tables stacked vertically."""
        ws = wb.create_sheet("Weekly Roster", 0)
        
        # Day names for header
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        
        # Shift display info (IST primary, EST overlap for US business hours)
        # IST is 10:30 hours ahead of EST
        shift_info = {
            ShiftType.SHIFT_1: {
                'name': "1st Shift (IST: 1:30 PM – 9:30 PM | EST: 3:00 AM – 11:00 AM)",
                'color': 'FFFFFF00'  # Yellow
            },
            ShiftType.SHIFT_2: {
                'name': "2nd Shift (IST: 9:30 PM – 5:30 AM | EST: 11:00 AM – 7:00 PM)",
                'color': 'FFFFFF00'  # Yellow
            },
            ShiftType.SHIFT_3: {
                'name': "3rd Shift (IST: 5:30 AM – 1:30 PM | EST: 7:00 PM – 3:00 AM)",
                'color': 'FFFFFF00'  # Yellow
            }
        }
        
        # Build weekly pattern from first week of schedule
        # Get which employees work on which day of week for each shift
        shift_weekly_schedule = {
            ShiftType.SHIFT_1: {day: [] for day in range(7)},
            ShiftType.SHIFT_2: {day: [] for day in range(7)},
            ShiftType.SHIFT_3: {day: [] for day in range(7)}
        }
        
        # Use the first 7 days (or available days) to determine pattern
        for day_assignment in roster.schedule[:7]:
            day_idx = day_assignment.date.weekday()
            
            for shift_type in [ShiftType.SHIFT_1, ShiftType.SHIFT_2, ShiftType.SHIFT_3]:
                shift_assignment = day_assignment.get_shift(shift_type)
                employees_working = shift_assignment.get_all_assigned()
                
                for emp in employees_working:
                    if emp.name not in [e.name for e in shift_weekly_schedule[shift_type][day_idx]]:
                        shift_weekly_schedule[shift_type][day_idx].append(emp)
        
        current_row = 1
        
        # Create table for each shift
        for shift_type in [ShiftType.SHIFT_1, ShiftType.SHIFT_2, ShiftType.SHIFT_3]:
            info = shift_info[shift_type]
            
            # Row: Shift title header (merged, yellow background)
            ws.cell(row=current_row, column=1, value=info['name'])
            ws.merge_cells(start_row=current_row, start_column=1, 
                          end_row=current_row, end_column=7)
            
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color=info['color'], 
                                        end_color=info['color'], 
                                        fill_type='solid')
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
            
            current_row += 1
            
            # Row: Day headers (Mon, Tue, Wed, Thu, Fri, Sat, Sun)
            for col, day in enumerate(days, 1):
                cell = ws.cell(row=current_row, column=col, value=day)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                
                # Weekend columns - light blue background
                if col >= 6:  # Sat, Sun
                    cell.fill = PatternFill(start_color='FFE6F2FF', 
                                           end_color='FFE6F2FF', 
                                           fill_type='solid')
            
            current_row += 1
            
            # Get employees on this shift (no hierarchy sorting - shuffled display)
            shift_employees = [emp for emp in roster.employees 
                              if emp.assigned_shift == shift_type]
            # Don't sort by seniority - keep as-is for random distribution
            
            # Create a row for each employee
            for emp in shift_employees:
                for col, day_idx in enumerate(range(7), 1):
                    # Check if employee works on this day
                    # Use weekly_pattern if available, else check schedule
                    if emp.weekly_pattern:
                        working = emp.weekly_pattern.get(day_idx, True)
                    else:
                        # Fallback: check if emp is in this day's schedule
                        working = emp in shift_weekly_schedule[shift_type].get(day_idx, [])
                    
                    cell = ws.cell(row=current_row, column=col, 
                                  value=emp.name if working else "")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
                    
                    # Alternate row background for readability
                    if (shift_employees.index(emp) % 2) == 1:
                        cell.fill = PatternFill(start_color='FFF5F5F5', 
                                               end_color='FFF5F5F5', 
                                               fill_type='solid')
                
                current_row += 1
            
            # Add empty row between shift tables
            current_row += 1
        
        # Adjust column widths - all columns should fit employee names
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze the first row
        ws.freeze_panes = 'A2'
    
    def _create_individual_sheet(self, wb: Workbook, roster: Roster):
        """Create individual employee schedules sheet."""
        ws = wb.create_sheet("Individual Schedules", 1)
        
        # Headers
        headers = ["Employee", "Seniority", "Assigned Shift", "Working Days", 
                   "Days Off", "Weekend Shifts", "Night Shifts", "Leaves"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color=self.COLORS['header_font'])
            cell.fill = PatternFill(start_color=self.COLORS['header_bg'],
                                   end_color=self.COLORS['header_bg'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Data rows
        stats = roster.statistics
        if not stats:
            roster.calculate_statistics()
            stats = roster.statistics
        
        row = 2
        for emp in sorted(roster.employees, key=lambda e: e.seniority_rank):
            emp_stats = stats.employee_stats.get(emp.name, {})
            
            shift_name = ""
            if emp.assigned_shift:
                shift_names = {
                    ShiftType.SHIFT_1: "Shift 1 (IST: 1:30 PM - 9:30 PM)",
                    ShiftType.SHIFT_2: "Shift 2 (IST: 9:30 PM - 5:30 AM)",
                    ShiftType.SHIFT_3: "Shift 3 (IST: 5:30 AM - 1:30 PM)"
                }
                shift_name = shift_names.get(emp.assigned_shift, "")
            
            ws.cell(row=row, column=1, value=emp.name)
            ws.cell(row=row, column=2, value=emp.seniority_rank)
            ws.cell(row=row, column=3, value=shift_name)
            ws.cell(row=row, column=4, value=emp_stats.get("working_days", 0))
            ws.cell(row=row, column=5, value=emp_stats.get("days_off", 0))
            ws.cell(row=row, column=6, value=emp_stats.get("weekend_shifts", 0))
            ws.cell(row=row, column=7, value=emp_stats.get("night_shifts", 0))
            ws.cell(row=row, column=8, value=len(emp.leaves))
            
            # Borders and alignment
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        for col in range(4, 9):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_statistics_sheet(self, wb: Workbook, roster: Roster):
        """Create statistics and fairness analysis sheet."""
        ws = wb.create_sheet("Statistics & Fairness", 2)
        
        stats = roster.statistics
        if not stats:
            roster.calculate_statistics()
            stats = roster.statistics
        
        # Title
        ws.cell(row=1, column=1, value="Roster Statistics & Fairness Analysis")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # General stats
        ws.cell(row=3, column=1, value="General Statistics")
        ws.cell(row=3, column=1).font = Font(bold=True, size=12)
        
        general_stats = [
            ("Month/Year", f"{roster.get_month_name()} {roster.year}"),
            ("Total Days", roster.get_total_days()),
            ("Total Team Members", len(roster.employees)),
            ("Total Shifts Scheduled", stats.total_shifts_scheduled),
            ("Coverage Requirements Met", "YES" if stats.all_coverage_met else "NO"),
            ("All Leaves Accommodated", "YES" if stats.all_leaves_accommodated else "NO"),
        ]
        
        row = 4
        for label, value in general_stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Fairness metrics
        row += 1
        ws.cell(row=row, column=1, value="Fairness Metrics")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        fairness_stats = [
            ("Average Working Days", f"{stats.avg_working_days:.1f}"),
            ("Average Weekend Shifts", f"{stats.avg_weekend_shifts:.1f}"),
            ("Weekend Shifts Range", f"{stats.min_weekend_shifts} - {stats.max_weekend_shifts}"),
            ("Average Night Shifts", f"{stats.avg_night_shifts:.1f}"),
            ("Night Shifts Range", f"{stats.min_night_shifts} - {stats.max_night_shifts}"),
        ]
        
        for label, value in fairness_stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Warnings
        if stats.warnings:
            row += 1
            ws.cell(row=row, column=1, value="Warnings")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color='FFFF0000')
            row += 1
            
            for warning in stats.warnings:
                ws.cell(row=row, column=1, value=f"⚠ {warning}")
                row += 1
        
        # Shift distribution summary
        row += 1
        ws.cell(row=row, column=1, value="Shift Distribution")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        shift_counts = {ShiftType.SHIFT_1: 0, ShiftType.SHIFT_2: 0, ShiftType.SHIFT_3: 0}
        for emp in roster.employees:
            if emp.assigned_shift:
                shift_counts[emp.assigned_shift] += 1
        
        shift_labels = {
            ShiftType.SHIFT_1: "Shift 1 (Day)",
            ShiftType.SHIFT_2: "Shift 2 (Evening)",
            ShiftType.SHIFT_3: "Shift 3 (Night)"
        }
        
        for shift_type, count in shift_counts.items():
            ws.cell(row=row, column=1, value=shift_labels[shift_type])
            ws.cell(row=row, column=2, value=f"{count} employees")
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def export_to_csv(self, roster: Roster) -> str:
        """Export roster to CSV file."""
        filename = self._get_filename(roster, "csv")
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', newline='') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                "Date", "Day", 
                "Shift1_Lead", "Shift1_Backup", "Shift1_Members",
                "Shift2_Lead", "Shift2_Backup", "Shift2_Members",
                "Shift3_Lead", "Shift3_Backup", "Shift3_Members"
            ])
            
            # Data
            for day in roster.schedule:
                s1 = day.shift_1
                s2 = day.shift_2
                s3 = day.shift_3
                
                writer.writerow([
                    day.date.strftime("%Y-%m-%d"),
                    day.day_name,
                    s1.lead.name if s1.lead else "",
                    s1.backup.name if s1.backup else "",
                    ";".join([m.name for m in s1.members]),
                    s2.lead.name if s2.lead else "",
                    s2.backup.name if s2.backup else "",
                    ";".join([m.name for m in s2.members]),
                    s3.lead.name if s3.lead else "",
                    s3.backup.name if s3.backup else "",
                    ";".join([m.name for m in s3.members])
                ])
        
        return str(filepath)
